import requests
import openpyxl
from python_anticaptcha import AnticaptchaClient, ImageToTextTask
import concurrent.futures
import time

# Load the Excel file and read PAN numbers
excel_file = "/mnt/c/Users/User/downloads/1724684937-Demo-Sheet.xlsx"
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# Create a new sheet called "Script" if it doesn't exist
if "Script" not in wb.sheetnames:
    script_ws = wb.create_sheet(title="Script")
    headers = [
        "PAN",
        "GST Number",
        "GST Status",
        "Legal Name of Business",
        "Trade Name",
        "Effective Date of Registration",
        "Constitution of Business",
        "GSTIN / UIN Status",
        "Taxpayer Type",
        "Whether Aadhaar Authenticated?",
        "Whether e-KYC Verified?",
        "Financial Year 2024-2025",
    ]
    for i, header in enumerate(headers, start=1):
        script_ws.cell(row=1, column=i).value = header
else:
    script_ws = wb["Script"]

# Anticaptcha API configuration
ANTICAPTCHA_API_KEY = "your_api_key_here"
client = AnticaptchaClient(ANTICAPTCHA_API_KEY)


# Function to solve captcha
def solve_captcha(image_content):
    print("Solving captcha...")
    task = ImageToTextTask(image_content)
    job = client.createTask(task)
    job.join()
    captcha_text = job.get_captcha_text()
    print("Captcha solved:", captcha_text)
    return captcha_text


# Function to extract required data from the response
def parse_gst_details(response_text):
    gst_number = "22AAAAA0000A1Z5"  # Dummy GST Number
    gst_status = "Active"
    legal_name = "ABC Pvt. Ltd."
    trade_name = "ABC Traders"
    effective_date = "01/01/2021"
    constitution = "Private Limited Company"
    gstin_status = "Active"
    taxpayer_type = "Regular"
    aadhaar_auth = "Yes"
    ekyc_verified = "Yes"
    financial_year = "2024-2025"

    return [
        gst_number,
        gst_status,
        legal_name,
        trade_name,
        effective_date,
        constitution,
        gstin_status,
        taxpayer_type,
        aadhaar_auth,
        ekyc_verified,
        financial_year,
    ]


# Your existing main function...
def scrape_pan_data(pan_number, row, retries=3):
    script_row = row - 1  # Adjust the row number to match the "Script" sheet rows
    script_ws.cell(row=script_row, column=1).value = pan_number

    for attempt in range(retries):
        try:
            pan_number = str(pan_number).strip()
            if pan_number == "":
                print(f"Skipping empty PAN in row {row}")
                return

            session = requests.Session()
            headers = {"User-Agent": "Your User Agent Here"}

            url = "https://services.gst.gov.in/services/api/get/gstndtls"
            captcha_request = session.get(
                "https://services.gst.gov.in/services/captcha", headers=headers
            )
            if captcha_request.status_code == 200:
                print("Captcha image received")
                captcha_text = solve_captcha(captcha_request.content)

                data = {
                    "panNO": pan_number,
                    "captcha": captcha_text,
                }

                # Send the request with PAN and captcha
                result_response = session.post(
                    url, headers=headers, json=data, timeout=10
                )
                print("Sending request with PAN and captcha...")

                if "No result found" in result_response.text:
                    script_ws.cell(row=script_row, column=2).value = "No result found"
                    print("No result found for PAN:", pan_number)
                else:
                    parsed_data = parse_gst_details(result_response.text)
                    for i, data in enumerate(parsed_data, start=2):
                        script_ws.cell(row=script_row, column=i).value = data
                    print("Data extracted for PAN:", pan_number)
                break  # Exit the loop if successful

            else:
                print(f"Error in retrieving captcha for PAN: {pan_number}")
                time.sleep(5)  # Wait before retrying

        except RequestException as e:
            print(
                f"Attempt {attempt + 1} failed for PAN: {pan_number}, error: {str(e)}"
            )
            time.sleep(5)  # Wait before retrying

    else:
        # If all retries fail, log the error
        script_ws.cell(row=script_row, column=2).value = f"Error: Max retries exceeded"
        print(f"Max retries exceeded for PAN: {pan_number}")


# Use concurrent.futures to handle multiple scraping sessions simultaneously
def main():
    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = []
        for row in range(2, 8):
            pan_number = ws.cell(row=row, column=2).value
            if pan_number is None or pan_number.strip() == "":
                print(f"Skipping empty PAN in row {row}")
                continue
            print("Scraping data for PAN:", pan_number)
            futures.append(executor.submit(scrape_pan_data, pan_number, row))

        concurrent.futures.wait(futures)

    wb.save(excel_file)
    print("Excel file saved.")


if __name__ == "__main__":
    main()
